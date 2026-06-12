"""
Diagnostic script - run with: python test_xls_parse.py <path_to_xls_file>
"""
import sys
import re
import xlrd
import openpyxl
import io

def _clean(h):
    if not h:
        return ""
    c = str(h).upper()
    c = re.sub(r"\([^)]*\)", "", c)
    c = re.sub(r"[^A-Z0-9]", "", c)
    return c

if len(sys.argv) < 2:
    print("Usage: python test_xls_parse.py <file.xls or file.xlsx>")
    sys.exit(1)

path = sys.argv[1]
file_bytes = open(path, "rb").read()

print(f"\n=== File: {path}, size: {len(file_bytes)} bytes ===\n")

if path.lower().endswith(".xls"):
    try:
        wb = xlrd.open_workbook(file_contents=file_bytes)
        print(f"xlrd opened OK. Sheets: {wb.sheet_names()}")
        for sname in wb.sheet_names():
            ws = wb.sheet_by_name(sname)
            print(f"\n--- Sheet: '{sname}', rows={ws.nrows}, cols={ws.ncols} ---")
            for r in range(min(ws.nrows, 10)):
                row_vals = []
                for c in range(ws.ncols):
                    cell = ws.cell(r, c)
                    row_vals.append(repr(cell.value)[:30])
                print(f"  Row {r}: {row_vals}")
    except Exception as e:
        print(f"xlrd FAILED: {e}")
        # Try as xlsx
        try:
            wb2 = openpyxl.load_workbook(filename=io.BytesIO(file_bytes), data_only=True)
            print(f"\nopenpyxl can open it! Sheets: {wb2.sheetnames}")
            for sname in wb2.sheetnames:
                ws2 = wb2[sname]
                print(f"\n--- Sheet: '{sname}' ---")
                for i, row in enumerate(ws2.iter_rows(values_only=True)):
                    if i >= 10: break
                    print(f"  Row {i}: {[repr(str(v))[:30] if v is not None else 'None' for v in row]}")
        except Exception as e2:
            print(f"openpyxl also FAILED: {e2}")
else:
    wb = openpyxl.load_workbook(filename=io.BytesIO(file_bytes), data_only=True)
    print(f"openpyxl opened OK. Sheets: {wb.sheetnames}")
    for sname in wb.sheetnames:
        ws = wb[sname]
        print(f"\n--- Sheet: '{sname}' ---")
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= 10: break
            cleaned = [_clean(v) if v else "" for v in row]
            print(f"  Row {i}: {[repr(v)[:25] for v in row]}")
            print(f"  Cleaned: {cleaned}")
