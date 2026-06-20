import io
import re
from datetime import date, datetime
from typing import Any, Optional
import openpyxl
import xlrd
from openpyxl.styles import PatternFill, Font


COLUMNS = [
    "Sr",
    "AS PER",
    "GSTN",
    "Recipient Name",
    "State",
    "Pos",
    "Invoice Num",
    "date",
    "Invoice Value",
    "Taxable Value",
    "IGST",
    "CGST",
    "SGST"
]

# We need some flexibility for exact header match since they might have spaces or (₹)
def _clean_header(h: str) -> str:
    if not h:
        return ""
    cleaned = str(h).upper()
    cleaned = re.sub(r"\([^)]*\)", "", cleaned)
    cleaned = re.sub(r"[^A-Z0-9]", "", cleaned)
    return cleaned

def _map_header(raw_header: str) -> str | None:
    """
    Fuzzy mapping from any portal/software column header to our internal standard column.
    Handles all GST portal variations without requiring exact key matches.
    """
    c = _clean_header(raw_header)
    if not c:
        return None
    # GSTIN
    if "GSTIN" in c or c == "GSTN" or "UIN" in c:
        return "GSTN"
    # Supplier/Party name
    if "TRADELEGAL" in c or "TRADENAME" in c or "LEGALNAME" in c or "RECIPIENTNAME" in c or "PARTYNAME" in c or "SUPPLIER" in c:
        return "Recipient Name"
    # Invoice Number
    if ("INVOICE" in c or "VOUCHER" in c or "BILL" in c or "DOCUMENT" in c or "NOTE" in c) and ("NUM" in c or c.endswith("NO") or "NO" in c and "DATE" not in c):
        return "Invoice Num"
    # Invoice Date
    if ("INVOICE" in c or "VOUCHER" in c or "BILL" in c or "DOCUMENT" in c or "NOTE" in c) and "DATE" in c:
        return "date"
    if c == "DATE":
        return "date"
    # Invoice Value (total) — check before TAXABLE
    if ("INVOICE" in c or "VOUCHER" in c or "BILL" in c or "NOTE" in c or "GROSS" in c) and ("VALUE" in c or "AMOUNT" in c or "TOTAL" in c):
        return "Invoice Value"
    # Taxable Value
    if "TAXABLE" in c:
        return "Taxable Value"
    # IGST
    if c == "IGST" or ("INTEGRATED" in c and "TAX" in c):
        return "IGST"
    # CGST
    if c == "CGST" or ("CENTRAL" in c and "TAX" in c):
        return "CGST"
    # SGST
    if c == "SGST" or "STATEUT" in c or ("STATE" in c and "TAX" in c):
        return "SGST"
    # Place of Supply / Pos / State
    if "PLACEOFSUPPLY" in c or "PLACEOFORIGIN" in c:
        return "Pos"
    if c in ("POS", "STATE"):
        return "Pos"
    return None

def normalize_key_str(s: str | None) -> str:
    """Strips all non-alphanumeric characters and uppercases.
    Used ONLY for building match keys, not for display.
    e.g. INV/2024/001 and INV2024001 both become INV2024001.
    """
    if not s:
        return ""
    return re.sub(r"[^A-Z0-9]", "", str(s).upper())

def fuzzy_invoice_match(inv1: str, inv2: str) -> bool:
    """
    Checks if two invoice strings are essentially the same.
    Used to pair invoices that have minor prefix differences but matching amounts.
    """
    if not inv1 or not inv2:
        return False
    
    inv1_str = str(inv1).strip()
    inv2_str = str(inv2).strip()
    
    # 1. Exact match after normalization
    n1 = normalize_key_str(inv1_str)
    n2 = normalize_key_str(inv2_str)
    if n1 == n2:
        return True
        
    # Extract digit sequences
    d1_list = re.findall(r'\d+', inv1_str)
    d2_list = re.findall(r'\d+', inv2_str)
    
    if d1_list and d2_list:
        # 2. Extracted numeric digits are identical (ignores all letters and special chars)
        if "".join(d1_list) == "".join(d2_list):
            return True
            
        # 3. Last numeric sequence matches exactly (as integers to ignore leading zeros)
        try:
            if int(d1_list[-1]) == int(d2_list[-1]):
                return True
        except ValueError:
            pass

    return False

GST_STATE_MAP = {
    "01": "JAMMUANDKASHMIR",
    "02": "HIMACHALPRADESH",
    "03": "PUNJAB",
    "04": "CHANDIGARH",
    "05": "UTTARAKHAND",
    "06": "HARYANA",
    "07": "DELHI",
    "08": "RAJASTHAN",
    "09": "UTTARPRADESH",
    "10": "BIHAR",
    "11": "SIKKIM",
    "12": "ARUNACHALPRADESH",
    "13": "NAGALAND",
    "14": "MANIPUR",
    "15": "MIZORAM",
    "16": "TRIPURA",
    "17": "MEGHALAYA",
    "18": "ASSAM",
    "19": "WESTBENGAL",
    "20": "JHARKHAND",
    "21": "ODISHA",
    "22": "CHHATTISGARH",
    "23": "MADHYAPRADESH",
    "24": "GUJARAT",
    "25": "DAMANANDDIU",
    "26": "DADRAANDNAGARHAVELI",
    "27": "MAHARASHTRA",
    "28": "ANDHRAPRADESH",
    "29": "KARNATAKA",
    "30": "GOA",
    "31": "LAKSHADWEEP",
    "32": "KERALA",
    "33": "TAMILNADU",
    "34": "PUDUCHERRY",
    "35": "ANDAMANANDNICOBARISLANDS",
    "36": "TELANGANA",
    "37": "ANDHRAPRADESH",
    "38": "LADAKH",
}

def normalize_state(s: str | None) -> str:
    if not s:
        return ""
    # Handle pure state codes (e.g. "24" or 24)
    s_str = str(s).strip()
    if s_str.isdigit() and len(s_str) in (1, 2):
        code = s_str.zfill(2)
        if code in GST_STATE_MAP:
            return GST_STATE_MAP[code]
            
    # Remove leading digits and hyphens, e.g. "09-Uttar Pradesh" -> "Uttar Pradesh"
    cleaned = re.sub(r"^[\d\-]+", "", s_str)
    cleaned = re.sub(r"\s+", "", cleaned).upper()
    return cleaned

def _safe_float(val: Any) -> float:
    try:
        if val is None or str(val).strip() == "":
            return 0.0
        return float(val)
    except ValueError:
        return 0.0

def _safe_date_str(val: Any) -> str:
    if isinstance(val, datetime):
        return val.strftime("%d-%m-%Y")
    if isinstance(val, date):
        return val.strftime("%d-%m-%Y")
    if val:
        s = str(val).strip()
        # Parse ISO date from DB (e.g. 2026-05-26)
        if len(s) >= 10 and s[4] == '-' and s[7] == '-':
            try:
                return datetime.strptime(s[:10], "%Y-%m-%d").strftime("%d-%m-%Y")
            except Exception:
                pass
        
        s = s.replace("/", "-")
        for fmt in ("%d-%m-%Y", "%d-%b-%Y", "%d-%b-%y", "%d-%m-%y"):
            try:
                return datetime.strptime(s, fmt).strftime("%d-%m-%Y")
            except ValueError:
                pass
        return s
    return ""


def _parse_date_to_obj(date_str: str) -> Optional[date]:
    """Convert a dd-mm-yyyy or ISO date string to a date object. Returns None on failure."""
    if not date_str:
        return None
    for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d-%b-%Y", "%d-%b-%y", "%d-%m-%y"):
        try:
            return datetime.strptime(date_str.strip(), fmt).date()
        except ValueError:
            pass
    return None

def build_software_rows(vouchers: list[dict[str, Any]], expected_type: str) -> list[dict[str, Any]]:
    """
    Transforms DB voucher records into normalized 11-column dicts.
    expected_type is either "purchases" or "debit_notes".
    """
    rows = []
    for idx, v in enumerate(vouchers, start=1):
        # Determine Invoice Type mapped string
        inv_type = "Regular"
        if v.get("category") == "Debit Note":
            inv_type = "Debit Note"
        elif v.get("category") == "Credit Note":
            inv_type = "Credit Note"

        # Calculate Grand Total from party accounting line
        # Assuming the query returns 'party_amount' or we calculate it
        # The query should join inventory lines and aggregate taxes
        taxable = v.get("total_taxable", 0.0)
        igst = v.get("total_igst", 0.0)
        cgst = v.get("total_cgst", 0.0)
        sgst = v.get("total_sgst", 0.0)
        
        # In a real scenario, grand total may include roundings or other ledgers.
        # Here we use party_amount if provided, else sum of items
        grand_total = v.get("party_amount", taxable + igst + cgst + sgst)

        pos_str = v.get("state", "").upper()

        rows.append({
            "Sr": idx,
            "AS PER": "BOOKS",
            "GSTN": v.get("gstin", ""),
            "Recipient Name": v.get("party_name", ""),
            "State": pos_str,
            "Pos": pos_str,
            "Invoice Num": v.get("voucher_number", ""),
            "date": _safe_date_str(v.get("voucher_date")),
            "Invoice Value": grand_total,
            "Taxable Value": taxable,
            "IGST": igst,
            "CGST": cgst,
            "SGST": sgst,
        })
    return rows


def parse_gstr2a_excel(file_bytes: bytes, filename: str) -> list[dict[str, Any]]:
    """
    Parses uploaded GSTR-2A Excel and extracts rows matching our 11 columns across all sheets.
    Handles both .xlsx (via openpyxl) and .xls (via xlrd).
    """
    results = []
    is_old_xls = filename.lower().endswith('.xls') and not filename.lower().endswith('.xlsx')
    print(f"[GSTR2A DEBUG] Filename: {filename}, is_old_xls: {is_old_xls}, file_size: {len(file_bytes)} bytes")

    def sheet_iterator():
        if is_old_xls:
            print("[GSTR2A DEBUG] Using xlrd for .xls file")
            wb = xlrd.open_workbook(file_contents=file_bytes)
            print(f"[GSTR2A DEBUG] xlrd sheets: {wb.sheet_names()}")
            for sheet_name in wb.sheet_names():
                ws = wb.sheet_by_name(sheet_name)
                print(f"[GSTR2A DEBUG] Sheet '{sheet_name}': {ws.nrows} rows x {ws.ncols} cols")
                def row_iterator():
                    for r_idx in range(ws.nrows):
                        row_vals = []
                        for c_idx in range(ws.ncols):
                            cell = ws.cell(r_idx, c_idx)
                            val = cell.value
                            if cell.ctype == xlrd.XL_CELL_DATE:
                                try:
                                    dt_tuple = xlrd.xldate_as_tuple(val, wb.datemode)
                                    val = datetime(*dt_tuple)
                                except Exception:
                                    pass
                            row_vals.append(val.strip("'") if isinstance(val, str) else val)
                        yield row_vals
                yield sheet_name, row_iterator()
        else:
            print("[GSTR2A DEBUG] Using openpyxl for .xlsx file")
            wb = openpyxl.load_workbook(filename=io.BytesIO(file_bytes), data_only=True)
            print(f"[GSTR2A DEBUG] openpyxl sheets: {wb.sheetnames}")
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                print(f"[GSTR2A DEBUG] Sheet '{sheet_name}': {ws.max_row} rows x {ws.max_column} cols")
                def row_iterator():
                    for row in ws.iter_rows(values_only=True):
                        yield row
                yield sheet_name, row_iterator()

    for sheet_name, row_iter in sheet_iterator():
        rows = list(row_iter)
        header_last_row_idx = -1
        col_map = {}

        # Print first 10 rows for debugging
        print(f"[GSTR2A DEBUG] Sheet '{sheet_name}' first 10 rows:")
        for i, row in enumerate(rows[:10]):
            print(f"  Row {i}: {[str(c)[:40] if c is not None else None for c in row]}")

        # Find headers - scan first 20 rows because GST portal has multi-row headers
        for r_idx, row in enumerate(rows[:20]):
            if not any(row):
                continue

            row_map = {}
            for c_idx, cell_val in enumerate(row):
                mapped = _map_header(cell_val)
                if mapped:
                    row_map[c_idx] = mapped
                    print(f"[GSTR2A DEBUG]   Header mapped: col {c_idx} '{cell_val}' -> '{mapped}'")

            if len(row_map) >= 3:
                col_map = row_map
                header_last_row_idx = r_idx
                print(f"[GSTR2A DEBUG]   Found header row at index {r_idx} with {len(row_map)} mapped columns")
                break

        has_gstin = any(v == "GSTN" for v in col_map.values())
        has_inv = any(v == "Invoice Num" for v in col_map.values())
        print(f"[GSTR2A DEBUG] Sheet '{sheet_name}': col_map={col_map}, has_gstin={has_gstin}, has_inv={has_inv}, header_last_row_idx={header_last_row_idx}")

        if header_last_row_idx == -1 or not has_inv:
            print(f"[GSTR2A DEBUG] SKIPPING sheet '{sheet_name}' - no Invoice Num header found")
            continue  # Skip this sheet, couldn't find minimum required headers

        # Parse data rows
        data_row_count = 0
        for r_idx, row in enumerate(rows):
            if r_idx <= header_last_row_idx:
                continue
            if not any(row):
                continue

            row_data = {col: "" for col in COLUMNS}
            row_data["AS PER"] = "GSTN"
            for col in ["Invoice Value", "Taxable Value", "IGST", "CGST", "SGST"]:
                row_data[col] = 0.0

            for c_idx, standard_col in col_map.items():
                if c_idx >= len(row):
                    continue
                val = row[c_idx]
                if standard_col in ["Invoice Value", "Taxable Value", "IGST", "CGST", "SGST"]:
                    row_data[standard_col] = _safe_float(val)
                elif standard_col == "date":
                    row_data[standard_col] = _safe_date_str(val)
                else:
                    row_data[standard_col] = str(val).strip() if val is not None else ""

            pos_str = row_data["Pos"]
            if pos_str:
                pos_str = re.sub(r'^\d+\s*-\s*', '', pos_str).strip().upper()
                row_data["Pos"] = pos_str
                row_data["State"] = pos_str

            gstin = row_data["GSTN"]
            inv_no = row_data["Invoice Num"]
            if inv_no:
                # Tag the sheet type so reconcile() can handle B2BA amendments correctly
                sheet_upper = sheet_name.upper()
                if "B2BA" in sheet_upper or "CDNRA" in sheet_upper:
                    row_data["_sheet_type"] = "B2BA"
                else:
                    row_data["_sheet_type"] = "B2B"
                results.append(row_data)
                data_row_count += 1

        print(f"[GSTR2A DEBUG] Sheet '{sheet_name}': extracted {data_row_count} data rows")

    print(f"[GSTR2A DEBUG] TOTAL extracted rows: {len(results)}")
    if not results:
        raise ValueError("Could not extract any data rows. Please ensure the file has the standard columns like Invoice Number, etc.")

    return results



def reconcile(
    software_rows: list[dict[str, Any]], 
    gstr2a_rows: list[dict[str, Any]],
    taxable_tolerance: float = 1.0,
    from_date: Optional[date] = None,
    to_date: Optional[date] = None,
) -> dict[str, Any]:
    from collections import defaultdict
    warnings: list[str] = []

    def _check_deviations(s_row: dict, g_row: dict) -> list[str]:
        """Compare two matched rows and return list of field names that deviate."""
        devs = []

        # We no longer flag Invoice Num as a deviation if they are a fuzzy match.
        if not fuzzy_invoice_match(s_row["Invoice Num"], g_row["Invoice Num"]):
            devs.append("Invoice Num")

        if abs(s_row["Taxable Value"] - g_row["Taxable Value"]) > taxable_tolerance:
            devs.append("Taxable Value")

        tax_tolerance = 2.0
        for col in ["IGST", "CGST", "SGST"]:
            if abs(s_row[col] - g_row[col]) > tax_tolerance:
                devs.append(col)

        # Invoice Value difference bounded by max possible legitimate sub-deviations
        if abs(s_row["Invoice Value"] - g_row["Invoice Value"]) > (taxable_tolerance + tax_tolerance):
            devs.append("Invoice Value")

        # Tax type swap: one side IGST, other side CGST+SGST, but total is same
        s_igst, s_cgst, s_sgst = s_row["IGST"], s_row["CGST"], s_row["SGST"]
        g_igst, g_cgst, g_sgst = g_row["IGST"], g_row["CGST"], g_row["SGST"]
        
        if s_igst > 0 and g_igst == 0:
            if abs(s_igst - (g_cgst + g_sgst)) <= tax_tolerance:
                if "IGST" in devs: devs.remove("IGST")
                if "CGST" in devs: devs.remove("CGST")
                if "SGST" in devs: devs.remove("SGST")
            elif "Tax Type (IGST vs CGST+SGST)" not in devs:
                devs.append("Tax Type (IGST vs CGST+SGST)")
        elif g_igst > 0 and s_igst == 0:
            if abs(g_igst - (s_cgst + s_sgst)) <= tax_tolerance:
                if "IGST" in devs: devs.remove("IGST")
                if "CGST" in devs: devs.remove("CGST")
                if "SGST" in devs: devs.remove("SGST")
            elif "Tax Type (IGST vs CGST+SGST)" not in devs:
                devs.append("Tax Type (IGST vs CGST+SGST)")

        # Date
        if s_row["date"] != g_row["date"]:
            devs.append("date")

        # State / Place of Supply
        if normalize_state(s_row["State"]) != normalize_state(g_row["State"]):
            devs.append("State")

        return devs

    def get_group_key(r: dict) -> str:
        g = normalize_key_str(r.get("GSTN", ""))
        if not g:
            return "__MISSING__"
        return g

    # ── Step 1: Group software rows by normalized GSTIN ──────────────────────
    software_by_gstin: dict[str, list] = defaultdict(list)
    for r in software_rows:
        software_by_gstin[get_group_key(r)].append(r)

    # ── Step 2: Group GSTR-2A rows by GSTIN, resolve B2BA and duplicates ─────
    # gstr2a_by_gstin[gstin][inv_key] = row
    gstr2a_by_gstin: dict[str, dict] = defaultdict(dict)
    for r in gstr2a_rows:
        gstin_key = get_group_key(r)
        inv_key   = normalize_key_str(r["Invoice Num"])
        if inv_key in gstr2a_by_gstin[gstin_key]:
            existing = gstr2a_by_gstin[gstin_key][inv_key]
            if r.get("_sheet_type") == "B2BA":
                # Amendment replaces the original
                gstr2a_by_gstin[gstin_key][inv_key] = r
            else:
                # Plain duplicate — sum tax values
                existing["Taxable Value"] += r["Taxable Value"]
                existing["IGST"]          += r["IGST"]
                existing["CGST"]          += r["CGST"]
                existing["SGST"]          += r["SGST"]
                existing["Invoice Value"]  = max(existing["Invoice Value"], r["Invoice Value"])
        else:
            gstr2a_by_gstin[gstin_key][inv_key] = r

    # ── Step 2b: Detect outside-range portal rows (if date range provided) ─────
    outside_range_grouped: dict[str, list] = {}
    if from_date and to_date:
        filtered_gstr2a_rows = []
        for r in gstr2a_rows:
            row_date = _parse_date_to_obj(r.get("date", ""))
            if row_date and (row_date < from_date or row_date > to_date):
                gstin_key = get_group_key(r)
                r_copy = r.copy()
                r_copy["Remark"] = "Outside Date Range"
                outside_range_grouped.setdefault(gstin_key, []).append(r_copy)
            else:
                filtered_gstr2a_rows.append(r)
        gstr2a_rows = filtered_gstr2a_rows
        # Rebuild the gstr2a_by_gstin map with filtered rows
        gstr2a_by_gstin = defaultdict(dict)
        for r in gstr2a_rows:
            gstin_key = get_group_key(r)
            inv_key   = normalize_key_str(r["Invoice Num"])
            if inv_key in gstr2a_by_gstin[gstin_key]:
                existing = gstr2a_by_gstin[gstin_key][inv_key]
                if r.get("_sheet_type") == "B2BA":
                    gstr2a_by_gstin[gstin_key][inv_key] = r
                else:
                    existing["Taxable Value"] += r["Taxable Value"]
                    existing["IGST"]          += r["IGST"]
                    existing["CGST"]          += r["CGST"]
                    existing["SGST"]          += r["SGST"]
                    existing["Invoice Value"]  = max(existing["Invoice Value"], r["Invoice Value"])
            else:
                gstr2a_by_gstin[gstin_key][inv_key] = r

    # ── Step 3: Match invoice-by-invoice within each party ───────────────────
    matched_grouped:          dict[str, list] = {}
    partial_grouped:          dict[str, list] = {}
    not_at_site_grouped:      dict[str, list] = {}
    not_in_software_grouped:  dict[str, list] = {}

    all_gstins = set(software_by_gstin.keys()) | set(gstr2a_by_gstin.keys())

    def _amounts_match(s_r: dict, g_r: dict) -> bool:
        if abs(s_r["Taxable Value"] - g_r["Taxable Value"]) > taxable_tolerance:
            return False
        tax_tol = 2.0
        for col in ["IGST", "CGST", "SGST"]:
            if abs(s_r[col] - g_r[col]) > tax_tol:
                return False
        if abs(s_r["Invoice Value"] - g_r["Invoice Value"]) > (taxable_tolerance + tax_tol):
            return False
        return True

    for gstin in all_gstins:
        s_rows   = software_by_gstin.get(gstin, [])
        g_inv_map = gstr2a_by_gstin.get(gstin, {})

        # Build software invoice map; warn on duplicate invoice numbers
        s_inv_map: dict[str, dict] = {}
        for r in s_rows:
            inv_key = normalize_key_str(r["Invoice Num"])
            if inv_key in s_inv_map:
                warnings.append(
                    f"Duplicate in books ignored: GSTIN '{r['GSTN']}', Invoice '{r['Invoice Num']}'"
                )
            else:
                s_inv_map[inv_key] = r

        matched_s: set[str] = set()
        matched_g: set[str] = set()

        # Pass 1: Exact normalized match
        for inv_key, g_row in g_inv_map.items():
            if inv_key in s_inv_map:
                s_row = s_inv_map[inv_key]
                matched_s.add(inv_key)
                matched_g.add(inv_key)

                deviations = _check_deviations(s_row, g_row)

                if not deviations:
                    matched_grouped.setdefault(gstin, []).append(s_row)
                else:
                    partial_grouped.setdefault(gstin, []).append({
                        "software_row": s_row,
                        "portal_row":   g_row,
                        "deviations":   deviations,
                    })

        # Pass 2: Fuzzy match for remaining unmatched
        unmatched_s_keys = [k for k in s_inv_map.keys() if k not in matched_s]
        unmatched_g_keys = [k for k in g_inv_map.keys() if k not in matched_g]
        
        for g_key in unmatched_g_keys:
            g_row = g_inv_map[g_key]
            best_s_key = None
            
            for s_key in unmatched_s_keys:
                if s_key in matched_s:
                    continue
                s_row = s_inv_map[s_key]
                
                # Check if fuzzy invoice match and amounts match
                if fuzzy_invoice_match(s_row["Invoice Num"], g_row["Invoice Num"]) and _amounts_match(s_row, g_row):
                    best_s_key = s_key
                    break
            
            if best_s_key:
                s_row = s_inv_map[best_s_key]
                matched_s.add(best_s_key)
                matched_g.add(g_key)
                
                deviations = _check_deviations(s_row, g_row)
                if not deviations:
                    matched_grouped.setdefault(gstin, []).append(s_row)
                else:
                    partial_grouped.setdefault(gstin, []).append({
                        "software_row": s_row,
                        "portal_row":   g_row,
                        "deviations":   deviations,
                    })

        # Not in Software — portal invoices with no books match
        for inv_key, g_row in g_inv_map.items():
            if inv_key not in matched_g:
                r_copy = g_row.copy()
                r_copy["Remark"] = "Not in Software"
                not_in_software_grouped.setdefault(gstin, []).append(r_copy)

        # Not at Site — books invoices with no portal match
        for inv_key, s_row in s_inv_map.items():
            if inv_key not in matched_s:
                r_copy = s_row.copy()
                r_copy["Remark"] = "Not at Site"
                not_at_site_grouped.setdefault(gstin, []).append(r_copy)

    # ── Step 4: Summary counts ────────────────────────────────────────────────
    matched_count        = sum(len(v) for v in matched_grouped.values())
    partial_count        = sum(len(v) for v in partial_grouped.values())
    not_at_site_count    = sum(len(v) for v in not_at_site_grouped.values())
    not_in_software_count = sum(len(v) for v in not_in_software_grouped.values())
    outside_range_count  = sum(len(v) for v in outside_range_grouped.values())

    return {
        "matched_grouped":         matched_grouped,
        "partial_grouped":         partial_grouped,
        "not_at_site_grouped":     not_at_site_grouped,
        "not_in_software_grouped": not_in_software_grouped,
        "outside_range_grouped":   outside_range_grouped,
        "warnings":                warnings,
        "summary": {
            "matched":          matched_count,
            "partially_matched": partial_count,
            "not_at_site":      not_at_site_count,
            "not_in_software":  not_in_software_count,
            "outside_range":    outside_range_count,
        },
    }


def build_result_excel(reconciliation_result: dict[str, Any]) -> bytes:
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    header_font  = Font(bold=True)
    party_font   = Font(bold=True, color="FFFFFF")
    party_fill   = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    books_fill   = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
    portal_fill  = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")
    red_fill     = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

    summary      = reconciliation_result["summary"]

    def _party_name_from_rows(rows: list, key: str = "Recipient Name") -> str:
        """Pull display name from the first row in a group."""
        if not rows:
            return ""
        row = rows[0]
        if isinstance(row, dict) and "software_row" in row:
            row = row["software_row"]
        return str(row.get(key, "") or "")

    def _write_party_header(ws, current_row: int, gstin: str, party_name: str,
                            n_cols: int, count: int) -> int:
        """Write a full-width merged party header row. Returns next row index."""
        if not gstin or str(gstin) == "__MISSING__":
            label = f"  MISSING GST NUM  ·  {count} invoice(s)"
        else:
            label = f"  {party_name}  —  {gstin}  ·  {count} invoice(s)"
        ws.cell(row=current_row, column=1, value=label)
        ws.cell(row=current_row, column=1).font  = party_font
        ws.cell(row=current_row, column=1).fill  = party_fill
        if n_cols > 1:
            ws.merge_cells(start_row=current_row, start_column=1,
                           end_row=current_row, end_column=n_cols)
        return current_row + 1

    def _autofit(ws):
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    # ── 0. Summary ────────────────────────────────────────────────────────────
    ws_sum = wb.create_sheet("Summary", 0)
    ws_sum.append(["Reconciliation Summary"])
    ws_sum.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws_sum.append([])
    ws_sum.append(["Matched",           summary["matched"]])
    ws_sum.append(["Partially Matched", summary["partially_matched"]])
    ws_sum.append(["Not at Site",       summary["not_at_site"]])
    ws_sum.append(["Not in Software",   summary["not_in_software"]])
    for r in range(3, 7):
        ws_sum.cell(row=r, column=1).font = header_font

    # ── 1. Matched ────────────────────────────────────────────────────────────
    ws_m = wb.create_sheet("Matched")
    ws_m.append(COLUMNS)
    for c in range(1, len(COLUMNS) + 1):
        ws_m.cell(row=1, column=c).font = header_font
    cur = 2
    sr  = 1
    for gstin, rows in reconciliation_result["matched_grouped"].items():
        party_name = _party_name_from_rows(rows)
        cur = _write_party_header(ws_m, cur, gstin, party_name, len(COLUMNS), len(rows))
        for row_data in rows:
            rd = row_data.copy()
            rd["Sr"] = sr
            ws_m.append([rd.get(c, "") for c in COLUMNS])
            sr  += 1
            cur += 1
        ws_m.append([])   # blank separator between parties
        cur += 1
    _autofit(ws_m)

    # ── 2. Partially Matched ──────────────────────────────────────────────────
    ws_p = wb.create_sheet("Partially Matched")
    ph = ["Source"] + COLUMNS + ["Deviations"]
    ws_p.append(ph)
    for c in range(1, len(ph) + 1):
        ws_p.cell(row=1, column=c).font = header_font
    cur = 2
    sr  = 1
    for gstin, pairs in reconciliation_result["partial_grouped"].items():
        party_name = _party_name_from_rows(pairs)
        n_devs = sum(len(p["deviations"]) for p in pairs)
        label  = f"  {party_name}  —  {gstin}  ·  {len(pairs)} invoice(s)  ·  {n_devs} deviation(s)"
        ws_p.cell(row=cur, column=1, value=label).font = party_font
        ws_p.cell(row=cur, column=1).fill = party_fill
        ws_p.merge_cells(start_row=cur, start_column=1,
                         end_row=cur, end_column=len(ph))
        cur += 1

        for p in pairs:
            s_row = p["software_row"].copy()
            g_row = p["portal_row"]
            devs  = p["deviations"]
            dev_text = "  |  ".join(
                f"{d}: Books={s_row.get(d, '')} / Portal={g_row.get(d, '')}"
                for d in devs
            )

            # Books row
            s_row["Sr"] = sr
            books_vals  = ["Books"] + [s_row.get(c, "") for c in COLUMNS] + [dev_text]
            for c_idx, val in enumerate(books_vals, start=1):
                cell = ws_p.cell(row=cur, column=c_idx, value=val)
                cell.fill = books_fill
                col_name  = COLUMNS[c_idx - 2] if 2 <= c_idx <= len(COLUMNS) + 1 else ""
                if col_name in devs:
                    cell.fill = red_fill
            cur += 1

            # Portal row
            g_row_copy = g_row.copy()
            g_row_copy["Sr"] = sr
            portal_vals = ["Portal"] + [g_row_copy.get(c, "") for c in COLUMNS] + [""]
            for c_idx, val in enumerate(portal_vals, start=1):
                cell = ws_p.cell(row=cur, column=c_idx, value=val)
                cell.fill = portal_fill
                col_name  = COLUMNS[c_idx - 2] if 2 <= c_idx <= len(COLUMNS) + 1 else ""
                if col_name in devs:
                    cell.fill = red_fill
            cur += 1
            sr  += 1

        ws_p.append([])   # blank separator
        cur += 1
    _autofit(ws_p)

    # ── 3. Not at Site ────────────────────────────────────────────────────────
    ws_nas = wb.create_sheet("Not at Site")
    nas_headers = COLUMNS + ["Remark"]
    ws_nas.append(nas_headers)
    for c in range(1, len(nas_headers) + 1):
        ws_nas.cell(row=1, column=c).font = header_font
    cur = 2
    sr  = 1
    for gstin, rows in reconciliation_result["not_at_site_grouped"].items():
        party_name = _party_name_from_rows(rows)
        cur = _write_party_header(ws_nas, cur, gstin, party_name, len(nas_headers), len(rows))
        for row_data in rows:
            rd = row_data.copy()
            rd["Sr"] = sr
            ws_nas.append([rd.get(c, "") for c in COLUMNS] + [rd.get("Remark", "")])
            sr  += 1
            cur += 1
        ws_nas.append([])
        cur += 1
    _autofit(ws_nas)

    # ── 4. Not in Software ────────────────────────────────────────────────────
    ws_nis = wb.create_sheet("Not in Software")
    ws_nis.append(nas_headers)
    for c in range(1, len(nas_headers) + 1):
        ws_nis.cell(row=1, column=c).font = header_font
    cur = 2
    sr  = 1
    for gstin, rows in reconciliation_result["not_in_software_grouped"].items():
        party_name = _party_name_from_rows(rows)
        cur = _write_party_header(ws_nis, cur, gstin, party_name, len(nas_headers), len(rows))
        for row_data in rows:
            rd = row_data.copy()
            rd["Sr"] = sr
            ws_nis.append([rd.get(c, "") for c in COLUMNS] + [rd.get("Remark", "")])
            sr  += 1
            cur += 1
        ws_nis.append([])
        cur += 1
    _autofit(ws_nis)

    # ── 5. Outside Date Range ─────────────────────────────────────────────────
    outside_range_grouped = reconciliation_result.get("outside_range_grouped", {})
    if outside_range_grouped:
        odr_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
        odr_font = Font(bold=True, color="856404")
        ws_odr = wb.create_sheet("Outside Date Range")
        odr_headers = COLUMNS + ["Remark"]
        ws_odr.append(odr_headers)
        for c in range(1, len(odr_headers) + 1):
            cell = ws_odr.cell(row=1, column=c)
            cell.font = header_font
            cell.fill = PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid")
        cur = 2
        sr  = 1
        for gstin, rows in outside_range_grouped.items():
            party_name = _party_name_from_rows(rows)
            cur = _write_party_header(ws_odr, cur, gstin, party_name, len(odr_headers), len(rows))
            for row_data in rows:
                rd = row_data.copy()
                rd["Sr"] = sr
                ws_odr.append([rd.get(c, "") for c in COLUMNS] + [rd.get("Remark", "")])
                for c_idx in range(1, len(odr_headers) + 1):
                    ws_odr.cell(row=cur, column=c_idx).fill = odr_fill
                sr  += 1
                cur += 1
            ws_odr.append([])
            cur += 1
        _autofit(ws_odr)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
