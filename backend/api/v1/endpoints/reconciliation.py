import io
import zipfile
from datetime import date
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, Query, status
from fastapi.responses import Response
from pydantic import BaseModel
from typing import Optional, List

from core.helpers import get_profile_context, resolve_target_firm_id
from core.security import get_verified_jwt
from core.supabase import supabase
from core.gstr2a_helpers import parse_gstr2a_excel, reconcile, build_software_rows, build_result_excel, COLUMNS, normalize_key_str
import openpyxl
from openpyxl.styles import Font

router = APIRouter()


@router.post("/debug-excel")
async def debug_excel(
    file: UploadFile = File(...),
    jwt: str = Depends(get_verified_jwt),
):
    """Debug endpoint: returns sheet names + first rows of every sheet + parsed row count."""
    import openpyxl, io as _io
    file_bytes = await file.read()
    wb = openpyxl.load_workbook(filename=_io.BytesIO(file_bytes), data_only=True)
    
    debug_info = {"sheets": []}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_preview = []
        for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if r_idx >= 8:  # show first 8 rows per sheet
                break
            rows_preview.append([str(c)[:80] if c is not None else None for c in row])
        debug_info["sheets"].append({
            "name": sheet_name,
            "first_rows": rows_preview
        })

    from core.gstr2a_helpers import parse_gstr2a_excel
    try:
        parsed = parse_gstr2a_excel(file_bytes)
        debug_info["parsed_row_count"] = len(parsed)
        debug_info["first_parsed_rows"] = parsed[:3]
    except Exception as e:
        debug_info["parse_error"] = str(e)

    return debug_info

def _fetch_purchase_vouchers(target_firm_id: str, from_date: date, to_date: date, match_type: str):
    """
    Fetch vouchers to reconcile. 
    match_type == "purchases" -> category='Purchase'
    match_type == "debit_notes" -> category in ('Debit Note', 'Credit Note')
    """
    query = (
        supabase.table("vouchers")
        .select("*")
        .eq("firm_id", target_firm_id)
        .eq("is_cancelled", False)
        .gte("voucher_date", str(from_date))
        .lte("voucher_date", str(to_date))
    )
    if match_type == "purchases":
        query = query.eq("category", "Purchase")
    else:
        query = query.in_("category", ["Debit Note", "Credit Note"])
        
    vouchers = query.execute().data or []
    if not vouchers:
        return []

    voucher_ids = [v["id"] for v in vouchers]
    
    party_ledger_ids = [v["party_ledger_id"] for v in vouchers if v.get("party_ledger_id")]
    parties_map = {}
    if party_ledger_ids:
        ledgers = (
            supabase.table("ledgers")
            .select("id, name")
            .in_("id", party_ledger_ids)
            .execute()
        ).data or []
        for l in ledgers:
            parties_map[l["id"]] = {"name": l["name"]}
            
        party_details = (
            supabase.table("ledger_party_details")
            .select("ledger_id, gstin, state")
            .in_("ledger_id", party_ledger_ids)
            .execute()
        ).data or []
        for pd in party_details:
            if pd["ledger_id"] in parties_map:
                parties_map[pd["ledger_id"]].update({
                    "gstin": pd.get("gstin", ""),
                    "state": pd.get("state", "")
                })

    inv_lines = (
        supabase.table("voucher_inventory_lines")
        .select("voucher_id, taxable_amount, igst_amount, cgst_amount, sgst_amount")
        .in_("voucher_id", voucher_ids)
        .execute()
    ).data or []
    
    inv_map = {}
    for line in inv_lines:
        vid = line["voucher_id"]
        if vid not in inv_map:
            inv_map[vid] = {"taxable": 0.0, "igst": 0.0, "cgst": 0.0, "sgst": 0.0}
        inv_map[vid]["taxable"] += float(line.get("taxable_amount") or 0.0)
        inv_map[vid]["igst"] += float(line.get("igst_amount") or 0.0)
        inv_map[vid]["cgst"] += float(line.get("cgst_amount") or 0.0)
        inv_map[vid]["sgst"] += float(line.get("sgst_amount") or 0.0)

    acc_lines = (
        supabase.table("voucher_accounting_lines")
        .select("voucher_id, ledger_id, debit_amount, credit_amount")
        .in_("voucher_id", voucher_ids)
        .execute()
    ).data or []

    acc_map = {}
    for line in acc_lines:
        vid = line["voucher_id"]
        lid = line["ledger_id"]
        if vid not in acc_map:
            acc_map[vid] = {}
        acc_map[vid][lid] = float(line.get("debit_amount") or 0.0) + float(line.get("credit_amount") or 0.0)

    result = []
    for v in vouchers:
        pid = v.get("party_ledger_id")
        party_info = parties_map.get(pid, {}) if pid else {}
        inv_info = inv_map.get(v["id"], {"taxable": 0.0, "igst": 0.0, "cgst": 0.0, "sgst": 0.0})
        party_amount = acc_map.get(v["id"], {}).get(pid, 0.0) if pid else 0.0

        result.append({
            "category": v["category"],
            "voucher_number": v["voucher_number"],
            "voucher_date": v["voucher_date"],
            "gstin": party_info.get("gstin", ""),
            "party_name": party_info.get("name", ""),
            "state": party_info.get("state", ""),
            "total_taxable": inv_info["taxable"],
            "total_igst": inv_info["igst"],
            "total_cgst": inv_info["cgst"],
            "total_sgst": inv_info["sgst"],
            "party_amount": party_amount
        })
    return result


@router.get("/export-software-data")
async def export_software_data(
    firm_id: str = Query(...),
    from_date: date = Query(...),
    to_date: date = Query(...),
    match_type: str = Query("purchases"),
    jwt: str = Depends(get_verified_jwt),
):
    profile = get_profile_context(jwt)
    target_firm_id = resolve_target_firm_id(profile, firm_id)

    raw_vouchers = _fetch_purchase_vouchers(target_firm_id, from_date, to_date, match_type)
    software_rows = build_software_rows(raw_vouchers, match_type)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Software Data"
    
    ws.append(COLUMNS)
    header_font = Font(bold=True)
    for col_idx in range(1, len(COLUMNS) + 1):
        ws.cell(row=1, column=col_idx).font = header_font

    for r in software_rows:
        ws.append([r[c] for c in COLUMNS])

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = min(max_length + 2, 50)

    out = io.BytesIO()
    wb.save(out)
    
    return Response(
        content=out.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=software_{match_type}_{from_date}_to_{to_date}.xlsx"}
    )


@router.post("/gstr2a")
async def reconcile_gstr2a(
    firm_id: str = Form(...),
    from_date: date = Form(...),
    to_date: date = Form(...),
    match_type: str = Form("purchases"),
    tolerance: float = Form(1.0),
    file: UploadFile = File(...),
    jwt: str = Depends(get_verified_jwt),
):
    profile = get_profile_context(jwt)
    target_firm_id = resolve_target_firm_id(profile, firm_id)

    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Invalid file type. Only Excel files are supported.")

    file_bytes = await file.read()
    try:
        gstr2a_rows = parse_gstr2a_excel(file_bytes, file.filename)
    except zipfile.BadZipFile:
        raise HTTPException(status_code=400, detail="The uploaded file is not a valid .xlsx file. (Note: Older .xls files or CSV files are not supported. Please save as an .xlsx Workbook).")
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse Excel: {e}")

    raw_vouchers = _fetch_purchase_vouchers(target_firm_id, from_date, to_date, match_type)
    software_rows = build_software_rows(raw_vouchers, match_type)

    result = reconcile(software_rows, gstr2a_rows, tolerance=tolerance)
    return result


@router.post("/gstr2a/download")
async def download_reconciliation_report(
    firm_id: str = Form(...),
    from_date: date = Form(...),
    to_date: date = Form(...),
    match_type: str = Form("purchases"),
    tolerance: float = Form(1.0),
    file: UploadFile = File(...),
    jwt: str = Depends(get_verified_jwt),
):
    profile = get_profile_context(jwt)
    target_firm_id = resolve_target_firm_id(profile, firm_id)

    file_bytes = await file.read()
    try:
        gstr2a_rows = parse_gstr2a_excel(file_bytes, file.filename)
    except zipfile.BadZipFile:
        raise HTTPException(status_code=400, detail="The uploaded file is not a valid .xlsx file. (Note: Older .xls files or CSV files are not supported. Please save as an .xlsx Workbook).")
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse Excel: {e}")

    raw_vouchers = _fetch_purchase_vouchers(target_firm_id, from_date, to_date, match_type)
    software_rows = build_software_rows(raw_vouchers, match_type)

    result = reconcile(software_rows, gstr2a_rows, tolerance=tolerance)
    excel_bytes = build_result_excel(result)

    return Response(
        content=excel_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=reconciliation_report.xlsx"}
    )

